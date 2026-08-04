"""
xlsx_generator.py
=================
Genera el archivo Excel ejecutivo con 6 hojas analíticas.
Usa openpyxl con formato profesional: paleta maroon, condicionales, gráficas.

Hojas:
  1. Resumen Ejecutivo  — KPIs + semáforo + 5 hallazgos
  2. Comparativo        — WoW / MoM / YoY / YTD
  3. Por Producto       — ranking + tendencia 12 semanas
  4. Por Cliente        — concentración + tendencia
  5. Regional           — ventas por estado
  6. Oportunidades      — hallazgos, tipo, impacto, recomendación
"""

import os
from typing import Optional, List, Dict
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule

import design_tokens as dt
from design_tokens import (
    BRAND_MAROON, HIGHLIGHT_CREAM, NEG_VALUE, POS_VALUE,
    SECTION_SUBTITLE, RULE_LINE, PAGE_BG,
    PRODUCT_ORDER, PRODUCT_DISPLAY_NAMES, PRODUCT_COLORS,
    CANAL_ORDER, CANAL_COLORS,
    fmt_currency, fmt_pct, fmt_int, fmt_ticket,
)
from data_processor import LocoDataProcessor


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    c = hex_color.lstrip("#")
    return PatternFill("solid", fgColor=c)


def _font(bold=False, color="000000", size=9, italic=False) -> Font:
    c = color.lstrip("#")
    return Font(bold=bold, color=c, size=size, italic=italic, name="Calibri")


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_thin() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


def _border_medium() -> Border:
    s = Side(style="medium", color="6E1E28")
    return Border(bottom=s)


FILL_MAROON  = _fill(BRAND_MAROON)
FILL_CREAM   = _fill(HIGHLIGHT_CREAM)
FILL_WHITE   = _fill("FFFFFF")
FILL_HEADER2 = _fill("F2F2F2")

FONT_HEADER  = _font(bold=True, color="FFFFFF", size=9)
FONT_TITLE   = _font(bold=True, color=BRAND_MAROON, size=14)
FONT_SECTION = _font(bold=True, color=BRAND_MAROON, size=11)
FONT_BODY    = _font(size=9)
FONT_BOLD    = _font(bold=True, size=9)
FONT_TOTAL   = _font(bold=True, color=BRAND_MAROON, size=9)
FONT_NEG     = _font(bold=True, color=NEG_VALUE, size=9)
FONT_SUB     = _font(italic=True, color=SECTION_SUBTITLE, size=8)

BORDER = _border_thin()

# Formatos de moneda: anteponen "$" y respetan separador de miles con comas.
# El "$" va entre comillas para que Excel lo trate como texto literal.
FMT_CURRENCY   = '"$"#,##0'
FMT_PCT        = '0%'
FMT_PCT1       = '0.0%'
FMT_PCT_SIGNED = '+0.0%;-0.0%;0.0%'   # variaciones: muestran el signo
FMT_INT        = '#,##0'              # conteos (botellas, cajas): sin "$"
FMT_DECIMAL    = '"$"#,##0.0'         # moneda con 1 decimal (ticket promedio)
FMT_PP         = '+0.0" pp";-0.0" pp";0.0" pp"'  # puntos porcentuales (margen)

TRAFFIC_GREEN  = "00B050"
TRAFFIC_YELLOW = "FFC000"
TRAFFIC_RED    = "FF0000"


def _set_col_width(ws, col_letter: str, width: float):
    ws.column_dimensions[col_letter].width = width


def _header_row(ws, row_idx: int, values: list, start_col: int = 1):
    """Escribe una fila de encabezado con estilo maroon."""
    for ci, val in enumerate(values, start_col):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill   = FILL_MAROON
        cell.font   = FONT_HEADER
        cell.alignment = _align("center")
        cell.border = BORDER


def _data_row(ws, row_idx: int, values: list, start_col: int = 1,
              bold=False, cream=False, neg_cols: list = None):
    """Escribe una fila de datos con estilo alternado."""
    fill = FILL_CREAM if cream else (_fill("F9F9F9") if row_idx % 2 == 0 else FILL_WHITE)
    for ci, val in enumerate(values, start_col):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill   = fill
        cell.font   = FONT_BOLD if bold else FONT_BODY
        cell.border = BORDER
        # Alineación: primera columna izquierda, resto derecha
        if ci == start_col:
            cell.alignment = _align("left")
        else:
            cell.alignment = _align("right")
        # Negativos en rojo
        if neg_cols and ci in neg_cols and isinstance(val, (int, float)) and val < 0:
            cell.font = FONT_NEG


def _traffic_light(ws, row_idx: int, col_idx: int, value: float, thresholds=(-5, 5)):
    """Semáforo: rojo < low, amarillo entre, verde > high."""
    cell = ws.cell(row=row_idx, column=col_idx)
    low, high = thresholds
    if isinstance(value, (int, float)):
        if value < low:
            color = TRAFFIC_RED
            emoji = "🔴"
        elif value > high:
            color = TRAFFIC_GREEN
            emoji = "🟢"
        else:
            color = TRAFFIC_YELLOW
            emoji = "🟡"
        cell.value = f"{emoji} {value:+.1f}%"
        cell.font  = _font(bold=True, color=color, size=9)
    cell.alignment = _align("center")


def _fix_chart_axes(chart):
    """
    Corrige los ejes que openpyxl deja mal configurados por defecto.

    openpyxl (3.1.x) crea ambos ejes con ``axPos='l'`` y sin el elemento
    ``<c:delete>``. Con ambos ejes "a la izquierda" Excel no puede dibujar el
    eje de categorías (horizontal) y solo se ven las líneas de cuadrícula. Aquí
    se fija la posición correcta según la orientación y se fuerza ``delete=0``
    (eje visible), de modo que los ejes se dibujen solos al abrir el archivo.
    """
    # BarChart con type="bar" es horizontal (categorías a la izquierda);
    # LineChart y columnas ("col") son verticales (categorías abajo).
    horizontal = getattr(chart, "type", None) == "bar"

    chart.x_axis.delete = False
    chart.y_axis.delete = False

    if horizontal:
        chart.x_axis.axPos = "l"   # eje de categorías a la izquierda
        chart.y_axis.axPos = "b"   # eje de valores abajo
    else:
        chart.x_axis.axPos = "b"   # eje de categorías abajo
        chart.y_axis.axPos = "l"   # eje de valores a la izquierda

    for ax in (chart.x_axis, chart.y_axis):
        ax.tickLblPos = "nextTo"
        ax.majorTickMark = "out"
        ax.minorTickMark = "none"


def _add_chart_with_note(ws, chart, anchor: str, span_cols: int = 8):
    """
    Agrega una gráfica al worksheet con los ejes ya corregidos para que Excel
    los dibuje automáticamente (ver :func:`_fix_chart_axes`).
    """
    _fix_chart_axes(chart)
    ws.add_chart(chart, anchor)


# ---------------------------------------------------------------------------
# Generador principal
# ---------------------------------------------------------------------------

class LocoReporteXLSX:
    """
    Genera el reporte Excel ejecutivo de Loco Tequila con 6 hojas analíticas.
    """

    def __init__(self, processor: LocoDataProcessor, output_path: str):
        self.proc        = processor
        self.output_path = output_path
        self.wb          = Workbook()
        self.wb.remove(self.wb.active)  # eliminar hoja vacía por defecto

    # ------------------------------------------------------------------
    # Generador principal
    # ------------------------------------------------------------------

    def generate(self):
        self._sheet_resumen()
        self._sheet_comparativo()
        self._sheet_productos()
        self._sheet_clientes()
        self._sheet_regional()
        self._sheet_canal()
        self._sheet_oportunidades()
        self.wb.save(self.output_path)
        print(f"[XLSX] Guardado: {self.output_path}")

    # ------------------------------------------------------------------
    # Hoja 1: Resumen Ejecutivo
    # ------------------------------------------------------------------

    def _sheet_resumen(self):
        ws = self.wb.create_sheet("Resumen Ejecutivo")
        ws.sheet_view.showGridLines = False

        r = self.proc.get_resumen_ejecutivo()
        sem = self.proc.semana
        anio = self.proc.anio

        # --- Título ---
        ws.merge_cells("A1:I1")
        ws["A1"] = f"Reporte Ejecutivo — Semana {sem:02d} · {anio}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = _align("left")
        ws["A1"].fill = _fill("FFF8F0")

        ws.merge_cells("A2:I2")
        ws["A2"] = f"Loco Tequila · Ventas y Margen · {anio}-W{sem:02d}"
        ws["A2"].font = FONT_SUB
        ws["A2"].fill = _fill("FFF8F0")

        # --- KPI Cards (fila 4-11) ---
        ws.merge_cells("A4:B4")
        ws["A4"] = "KPIs Clave — Semana Actual"
        ws["A4"].font = FONT_SECTION
        ws["A4"].fill = FILL_CREAM

        kpi_headers = ["Métrica", "Valor", "Plan", "Var vs Plan", "Var vs Plan %",
                       "Año Anterior", "Var vs Año Ant.", "Var vs Año Ant. %"]
        _header_row(ws, 5, kpi_headers, start_col=1)

        cur   = r["actual"]
        plan  = r["plan"]
        lasty = r["anio_anterior"]

        # (etiqueta, clave, tipo)  tipo: money | int | pct | money1
        metricas_kpi = [
            ("Ventas Netas (sin IVA)", "ventas_netas",     "money"),
            ("# Botellas Vendidas",    "botellas",          "int"),
            ("Cajas 9 Litros",         "cajas_9l",          "int"),
            ("Margen % Estimado",      "margen_pct",        "pct"),
            ("Ticket Promedio $",      "ticket_promedio",   "money1"),
        ]

        for ri, (met, key, kind) in enumerate(metricas_kpi, 6):
            val    = cur.get(key, 0)
            plan_v = plan.get(key, 0)
            ant_v  = lasty.get(key, 0)

            if kind == "pct":
                # Márgenes: el valor se guarda como fracción (formato %); la
                # variación se expresa en puntos porcentuales (pp) y en % relativo.
                val_cell, plan_cell, ant_cell = val / 100, plan_v / 100, ant_v / 100
                var_plan = val - plan_v            # pp (crudo)
                var_ant  = val - ant_v
                fmt_val, fmt_abs = FMT_PCT1, FMT_PP
            else:
                val_cell, plan_cell, ant_cell = val, plan_v, ant_v
                var_plan = val - plan_v
                var_ant  = val - ant_v
                fmt_val = FMT_CURRENCY if kind == "money" else (
                    FMT_DECIMAL if kind == "money1" else FMT_INT)
                fmt_abs = fmt_val

            var_plan_p = (var_plan / plan_v) if (kind != "pct" and plan_v) else (
                ((val - plan_v) / plan_v) if plan_v else None)
            var_ant_p  = (var_ant / ant_v) if (kind != "pct" and ant_v) else (
                ((val - ant_v) / ant_v) if ant_v else None)

            row_vals = [
                met,
                val_cell,
                plan_cell  if plan_v else "—",
                var_plan   if plan_v else "—",
                var_plan_p if var_plan_p is not None else "—",
                ant_cell   if ant_v else "—",
                var_ant    if ant_v else "—",
                var_ant_p  if var_ant_p is not None else "—",
            ]
            _data_row(ws, ri, row_vals, bold=(ri == 6), cream=(ri == 6),
                      neg_cols=[4, 7])

            # Formatos por celda (solo si la celda es numérica)
            fmt_map = {
                2: fmt_val, 3: fmt_val, 4: fmt_abs, 5: FMT_PCT_SIGNED,
                6: fmt_val, 7: fmt_abs, 8: FMT_PCT_SIGNED,
            }
            for col, fmt in fmt_map.items():
                cell = ws.cell(row=ri, column=col)
                # Aplica el formato a cualquier valor numérico (incluye numpy
                # int64/float64); solo se omite el marcador de texto "—".
                if not isinstance(cell.value, str) and cell.value is not None:
                    cell.number_format = fmt

        # --- Semáforos ---
        ws.merge_cells("A13:F13")
        ws["A13"] = "Semáforo de Desempeño"
        ws["A13"].font = FONT_SECTION
        ws["A13"].fill = FILL_CREAM

        semaforos = [
            ("WoW (vs semana anterior)",          r["vs_semana_anterior"]["pct"]),
            ("Semana vs Mismo Período Año Ant.",   r["vs_anio_anterior"]["pct"]),
            ("YTD vs Mismo Corte Año Anterior",    r["ytd_vs_ly"]["pct"]),
            ("vs Plan Semanal",                    r["vs_plan"]["pct"]),
        ]
        _header_row(ws, 14, ["Indicador", "Variación %", "Estado", "Lectura"], start_col=1)
        for ri, (label, pct) in enumerate(semaforos, 15):
            if pct < -10:
                lectura = "Atención: caída relevante — revisar causa raíz"
            elif pct < 0:
                lectura = "Ligera baja — monitorear"
            elif pct < 5:
                lectura = "Desempeño estable"
            elif pct < 15:
                lectura = "Crecimiento sólido — sostener estrategia"
            else:
                lectura = "Crecimiento fuerte — identificar driver y replicar"
            ws.cell(row=ri, column=1, value=label).font = FONT_BODY
            ws.cell(row=ri, column=1).alignment = _align("left")
            ws.cell(row=ri, column=2, value=f"{pct:+.1f}%").alignment = _align("right")
            ws.cell(row=ri, column=2).font = FONT_BODY
            _traffic_light(ws, ri, 3, pct)
            ws.cell(row=ri, column=4, value=lectura).font = FONT_SUB
            ws.cell(row=ri, column=4).alignment = _align("left", wrap=True)
            ws.cell(row=ri, column=4).border = BORDER
            for ci in range(1, 4):
                ws.cell(row=ri, column=ci).border = BORDER

        # --- 5 Hallazgos principales ---
        ws.merge_cells("A21:F21")
        ws["A21"] = "Hallazgos Principales"
        ws["A21"].font = FONT_SECTION
        ws["A21"].fill = FILL_CREAM

        hallazgos = self.proc.get_oportunidades_riesgos()
        _header_row(ws, 22, ["#", "Hallazgo", "Tipo", "Impacto", "Recomendación"], start_col=1)
        for ri, h in enumerate(hallazgos[:5], 23):
            tipo_color = TRAFFIC_RED if h["tipo"] == "Riesgo" else (
                TRAFFIC_GREEN if h["tipo"] == "Oportunidad" else TRAFFIC_YELLOW)
            vals = [ri - 22, h["hallazgo"], h["tipo"], h["impacto"], h["recomendacion"]]
            _data_row(ws, ri, vals)
            ws.cell(row=ri, column=3).font = _font(bold=True, color=tipo_color, size=9)

        # Anchos de columna (8 columnas de la tabla KPI)
        widths = [30, 16, 15, 15, 13, 16, 15, 13]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Hoja 2: Comparativo — 4 ventanas estándar + serie anual
    # ------------------------------------------------------------------

    def _sheet_comparativo(self):
        ws = self.wb.create_sheet("Comparativo Periodos")
        ws.sheet_view.showGridLines = False

        r   = self.proc.get_resumen_ejecutivo()
        sem = self.proc.semana
        anio = self.proc.anio
        ly   = anio - 1

        ws.merge_cells("A1:G1")
        ws["A1"] = f"Comparativo de Periodos — Semana {sem:02d} · {anio} — Ventas sin IVA (MXN)"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = _align("left")
        ws["A1"].fill = _fill("FFF8F0")

        # Nota metodológica
        ws.merge_cells("A2:G2")
        ws["A2"] = "Ventanas comparativas: WoW | Semana vs Mismo Período Año Anterior | YTD vs LY | Rolling 52 sem (si aplica)"
        ws["A2"].font = FONT_SUB
        ws["A2"].fill = _fill("FFF8F0")

        # ── BLOQUE A: WoW ──────────────────────────────────────────────
        sem_ant_v = r["actual"]["ventas_netas"] - r["vs_semana_anterior"]["abs"]
        self._write_comparison_block(
            ws, start_row=4,
            title=f"Bloque A — Semana a Semana (WoW): S{sem:02d} vs S{sem-1 if sem > 1 else 52:02d}",
            actual=r["actual"]["ventas_netas"],
            anterior=sem_ant_v,
            var_abs=r["vs_semana_anterior"]["abs"],
            var_pct=r["vs_semana_anterior"]["pct"],
            labels=[f"Semana {sem:02d}-{anio}", f"Semana {sem-1 if sem > 1 else 52:02d}-{anio if sem > 1 else ly}"],
            extra_kpis=r["actual"]
        )

        # ── BLOQUE B: YoY Semanal ──────────────────────────────────────
        yoy_sem_ant_v = r["actual"]["ventas_netas"] - r["vs_anio_anterior"]["abs"]
        self._write_comparison_block(
            ws, start_row=11,
            title=f"Bloque B — Semana vs Mismo Período Año Anterior: S{sem:02d}-{anio} vs S{sem:02d}-{ly}",
            actual=r["actual"]["ventas_netas"],
            anterior=yoy_sem_ant_v,
            var_abs=r["vs_anio_anterior"]["abs"],
            var_pct=r["vs_anio_anterior"]["pct"],
            labels=[f"S{sem:02d}-{anio}", f"S{sem:02d}-{ly}"]
        )

        # ── BLOQUE C: YTD vs LY ────────────────────────────────────────
        self._write_comparison_block(
            ws, start_row=18,
            title=f"Bloque C — Acumulado del Año (YTD): {anio} vs {ly} (hasta S{sem:02d})",
            actual=r["ytd_actual"]["ventas_netas"],
            anterior=r["ytd_actual"]["ventas_netas"] - r["ytd_vs_ly"]["abs"],
            var_abs=r["ytd_vs_ly"]["abs"],
            var_pct=r["ytd_vs_ly"]["pct"],
            labels=[f"YTD {anio} (S01–S{sem:02d})", f"YTD {ly} (S01–S{sem:02d})"]
        )

        # ── BLOQUE D: Rolling 52 semanas (condicional) ─────────────────
        row_after_blocks = 26
        rolling = self.proc.get_rolling_52()
        if rolling:
            self._write_comparison_block(
                ws, start_row=row_after_blocks,
                title=f"Bloque D — Rolling 52 Semanas ({rolling['semanas_incluidas']} sem con datos)",
                actual=rolling["ventas_rolling"],
                anterior=rolling["ventas_rolling_ly"],
                var_abs=rolling["var_abs"],
                var_pct=rolling["var_pct"],
                labels=[rolling["label_a"], rolling["label_b"]]
            )
            serie_start = row_after_blocks + 8
        else:
            # Nota explicativa si no hay datos suficientes
            ws.merge_cells(f"A{row_after_blocks}:G{row_after_blocks}")
            ws[f"A{row_after_blocks}"] = (
                "Bloque D — Rolling 52 Semanas: no disponible "
                "(se requieren al menos 40 semanas con datos históricos)"
            )
            ws[f"A{row_after_blocks}"].font = FONT_SUB
            ws[f"A{row_after_blocks}"].fill = _fill("F5F5F5")
            serie_start = row_after_blocks + 3

        # ── SERIE ANUAL COMPLETA con Promedio Móvil 4 sem ─────────────
        ws.merge_cells(f"A{serie_start}:G{serie_start}")
        ws[f"A{serie_start}"] = f"Serie Semanal Completa {anio} — Ventas Netas vs Año Anterior (S01–S{sem:02d})"
        ws[f"A{serie_start}"].font = FONT_SECTION
        ws[f"A{serie_start}"].fill = FILL_CREAM

        serie_anual = self.proc.get_serie_anual_semanal()
        plan_data   = self.proc.get_plan_semanal(n_semanas=sem + 2)
        plan_dict = {}
        if not plan_data.empty:
            plan_dict = dict(zip(plan_data["label"], plan_data["plan_venta_sin_impuestos"]))

        cols_hdr = [
            "Semana", f"Ventas {anio} $", f"Ventas {ly} $",
            "Plan $", "Var vs Plan $", "Var vs Plan %",
            "Promedio Móvil 4 sem"
        ]
        _header_row(ws, serie_start + 1, cols_hdr, start_col=1)

        chart_row_start = serie_start + 2
        for ri, (_, srow) in enumerate(serie_anual.iterrows(), serie_start + 2):
            lbl      = srow.get("label", "")
            v_cur    = float(srow.get("venta_actual", 0))
            v_ly     = float(srow.get("venta_anio_anterior", 0))
            mov_avg  = float(srow.get("promedio_movil_4s", 0))
            plan_v   = plan_dict.get(lbl, 0)
            var_a    = v_cur - plan_v
            var_p    = (var_a / plan_v * 100) if plan_v > 0 else 0
            cream    = (lbl == f"W{sem:02d}")
            _data_row(ws, ri, [lbl, v_cur, v_ly, plan_v, var_a, var_p / 100, mov_avg],
                      cream=cream, neg_cols=[5])
            for ci, fmt in [
                (2, FMT_CURRENCY), (3, FMT_CURRENCY), (4, FMT_CURRENCY),
                (5, FMT_CURRENCY), (6, FMT_PCT1), (7, FMT_CURRENCY)
            ]:
                ws.cell(row=ri, column=ci).number_format = fmt

        chart_row_end = serie_start + 1 + len(serie_anual)

        # ── Gráfica Line: ventas año actual + año anterior + promedio móvil ─
        from openpyxl.chart import LineChart, Reference as Ref
        try:
            lc = LineChart()
            lc.title = f"Tendencia Semanal — {anio} vs {ly} + Promedio Móvil 4 sem"
            lc.y_axis.title = "$MXN (sin IVA)"
            lc.x_axis.title = "Semana"
            lc.x_axis.tickLblPos = "nextTo"
            lc.y_axis.tickLblPos = "nextTo"
            lc.style = 10
            lc.width = 26
            lc.height = 13

            # Serie año actual (col 2)
            d1 = Ref(ws, min_col=2, min_row=serie_start + 1, max_row=chart_row_end)
            lc.add_data(d1, titles_from_data=True)
            # Serie año anterior (col 3)
            d2 = Ref(ws, min_col=3, min_row=serie_start + 1, max_row=chart_row_end)
            lc.add_data(d2, titles_from_data=True)
            # Promedio móvil (col 7)
            d3 = Ref(ws, min_col=7, min_row=serie_start + 1, max_row=chart_row_end)
            lc.add_data(d3, titles_from_data=True)

            cats = Ref(ws, min_col=1, min_row=serie_start + 2, max_row=chart_row_end)
            lc.set_categories(cats)

            # Estilos de líneas
            if lc.series:
                from openpyxl.chart.series import SeriesLabel
                from openpyxl.drawing.line import LineProperties
                lc.series[0].graphicalProperties.line.solidFill = BRAND_MAROON.lstrip("#")
                lc.series[0].graphicalProperties.line.width = 20000
                if len(lc.series) > 1:
                    lc.series[1].graphicalProperties.line.solidFill = "E7D6A6"
                    lc.series[1].graphicalProperties.line.width = 15000
                if len(lc.series) > 2:
                    lc.series[2].graphicalProperties.line.solidFill = "E23B2E"
                    lc.series[2].graphicalProperties.line.width = 12000
                    lc.series[2].graphicalProperties.line.dashDot = "dash"

            _add_chart_with_note(ws, lc, f"I{serie_start}")
        except Exception:
            pass  # gráfica opcional — no bloquea el reporte

        # Anchos
        for i, w in enumerate([12, 18, 18, 18, 18, 14, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_comparison_block(
        self, ws, start_row: int, title: str,
        actual: float, anterior: float,
        var_abs: float, var_pct: float,
        labels: list, extra_kpis: dict = None
    ):
        ws.merge_cells(f"A{start_row}:G{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = FONT_SECTION
        ws[f"A{start_row}"].fill = FILL_CREAM

        _header_row(ws, start_row + 1,
                    [labels[0], labels[1], "Variación $", "Variación %", "Tendencia"],
                    start_col=1)

        vals = [actual, anterior, var_abs, var_pct / 100,
                "▲" if var_abs >= 0 else "▼"]
        _data_row(ws, start_row + 2, vals, cream=True, neg_cols=[3])

        for ci, fmt in [(1, FMT_CURRENCY), (2, FMT_CURRENCY),
                         (3, FMT_CURRENCY), (4, FMT_PCT1)]:
            ws.cell(row=start_row + 2, column=ci).number_format = fmt

        crec_cell = ws.cell(row=start_row + 2, column=5)
        crec_cell.font = _font(bold=True,
                                color=TRAFFIC_GREEN if var_abs >= 0 else TRAFFIC_RED,
                                size=12)
        crec_cell.alignment = _align("center")

        # Fila extra con botellas y margen si se proporcionan KPIs
        if extra_kpis:
            ws.cell(row=start_row + 3, column=1,
                    value=f"Botellas: {extra_kpis.get('botellas', 0):,.0f}   "
                          f"Margen: ${extra_kpis.get('margen_pesos', 0):,.0f}   "
                          f"Ticket Prom.: ${extra_kpis.get('ticket_promedio', 0):,.2f}"
                    ).font = FONT_SUB
            ws.merge_cells(f"A{start_row + 3}:G{start_row + 3}")

    # ------------------------------------------------------------------
    # Hoja 3: Por Producto
    # ------------------------------------------------------------------

    def _sheet_productos(self):
        ws = self.wb.create_sheet("Por Producto")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:H1")
        ws["A1"] = "Análisis por Producto — Ranking y Tendencia"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = _align("left")
        ws["A1"].fill = _fill("FFF8F0")

        # Ranking anual
        ranking = self.proc.get_ranking_productos(mode="anual")
        ws.merge_cells("A3:H3")
        ws["A3"] = "Ranking por Volumen de Ventas (YTD)"
        ws["A3"].font = FONT_SECTION
        ws["A3"].fill = FILL_CREAM

        cols = ["Producto", "Ventas Netas $", "Botellas", "Cajas 9L",
                "Margen $", "Margen %", "Participación %"]
        _header_row(ws, 4, cols, start_col=1)

        for ri, (_, row) in enumerate(ranking.iterrows(), 5):
            cream = (ri == 5)  # top 1 resaltado
            vals = [
                PRODUCT_DISPLAY_NAMES.get(row["producto"], row["producto"]),
                row["ventas"],
                row["botellas"],
                row["cajas"],
                row["margen"],
                row["margen_pct"] / 100,
                row["participacion_pct"] / 100,
            ]
            _data_row(ws, ri, vals, cream=cream)
            fmts = [None, FMT_CURRENCY, FMT_INT, FMT_INT, FMT_CURRENCY, FMT_PCT1, FMT_PCT1]
            for ci, fmt in enumerate(fmts, 1):
                if fmt:
                    ws.cell(row=ri, column=ci).number_format = fmt

        # Fila total
        tot_row = len(ranking) + 5
        total_vals = ["Total",
                      ranking["ventas"].sum(),
                      ranking["botellas"].sum(),
                      ranking["cajas"].sum(),
                      ranking["margen"].sum(), "—", "100%"]
        _data_row(ws, tot_row, total_vals, bold=True, cream=True)
        ws.cell(row=tot_row, column=2).number_format = FMT_CURRENCY
        ws.cell(row=tot_row, column=3).number_format = FMT_INT

        # Tabla tendencia 12 semanas
        row_start = tot_row + 3
        ws.merge_cells(f"A{row_start}:M{row_start}")
        ws[f"A{row_start}"] = "Tendencia Semanal por Producto (últimas 12 semanas)"
        ws[f"A{row_start}"].font = FONT_SECTION
        ws[f"A{row_start}"].fill = FILL_CREAM

        serie = self.proc.get_serie_semanal(n_semanas=12)
        if not serie.empty:
            prods_presentes = [p for p in PRODUCT_ORDER if p in serie.columns]
            hdr = ["Semana"] + [PRODUCT_DISPLAY_NAMES.get(p, p) for p in prods_presentes] + ["Total"]
            _header_row(ws, row_start + 1, hdr)
            for ri, (_, srow) in enumerate(serie.iterrows(), row_start + 2):
                vals = [srow.get("label", "")]
                for p in prods_presentes:
                    vals.append(float(srow[p]))
                vals.append(float(srow.get("Total", 0)))
                cream = (srow.get("label", "") == f"W{self.proc.semana:02d}")
                _data_row(ws, ri, vals, cream=cream)
                for ci in range(2, len(vals) + 1):
                    ws.cell(row=ri, column=ci).number_format = FMT_CURRENCY

            chart_row_end = row_start + 1 + len(serie)

            # ── Gráfica de barras apiladas (tendencia semanal por producto) ──
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "stacked"
            chart.overlap = 100
            chart.title  = "Tendencia Semanal por Producto (barras apiladas)"
            chart.y_axis.title = "$MXN (sin IVA)"
            chart.x_axis.title = "Semana"
            chart.x_axis.tickLblPos = "nextTo"
            chart.y_axis.tickLblPos = "nextTo"
            chart.style = 10
            chart.width = 26
            chart.height = 13

            cats = Reference(ws, min_col=1, min_row=row_start + 2, max_row=chart_row_end)
            for idx_p, p in enumerate(prods_presentes):
                ci = 2 + idx_p
                data = Reference(ws, min_col=ci, min_row=row_start + 1, max_row=chart_row_end)
                series_obj = BarChart()
                series_obj.add_data(data, titles_from_data=True)
                s = series_obj.series[0]
                color_hex = PRODUCT_COLORS.get(p, "#999999").lstrip("#")
                s.graphicalProperties.solidFill = color_hex
                chart.append(s)
            chart.set_categories(cats)
            _add_chart_with_note(ws, chart, f"A{chart_row_end + 3}")

            # ── Tabla de clasificación de tendencia ──────────────────────────
            trend_start = chart_row_end + 20
            ws.merge_cells(f"A{trend_start}:F{trend_start}")
            ws[f"A{trend_start}"] = "Clasificación de Tendencia por Producto (1ª vs 2ª mitad de las últimas 12 semanas)"
            ws[f"A{trend_start}"].font = FONT_SECTION
            ws[f"A{trend_start}"].fill = FILL_CREAM

            _header_row(ws, trend_start + 1,
                        ["Producto", "Promedio 1ª Mitad $", "Promedio 2ª Mitad $",
                         "Variación %", "Tendencia"], start_col=1)

            # Calcular 1ª vs 2ª mitad
            n_rows_serie = len(serie)
            mid = n_rows_serie // 2
            for tr, p in enumerate(prods_presentes, trend_start + 2):
                if p not in serie.columns:
                    continue
                vals_prod = serie[p].tolist()
                primera   = sum(vals_prod[:mid]) / max(mid, 1)
                segunda   = sum(vals_prod[mid:]) / max(len(vals_prod) - mid, 1)
                var_tend  = ((segunda - primera) / primera * 100) if primera > 0 else 0
                if var_tend > 5:
                    tend_label = "Ascendente"
                    tend_fill  = _fill("1E7145")
                    tend_font  = _font(bold=True, color="FFFFFF", size=9)
                elif var_tend < -5:
                    tend_label = "Descendente"
                    tend_fill  = _fill("A6192E")
                    tend_font  = _font(bold=True, color="FFFFFF", size=9)
                else:
                    tend_label = "Estable"
                    tend_fill  = _fill("FFC000")
                    tend_font  = _font(bold=True, color="1F1F1F", size=9)

                row_vals = [
                    PRODUCT_DISPLAY_NAMES.get(p, p),
                    primera, segunda, var_tend / 100, tend_label
                ]
                _data_row(ws, tr, row_vals)
                ws.cell(row=tr, column=2).number_format = FMT_CURRENCY
                ws.cell(row=tr, column=3).number_format = FMT_CURRENCY
                ws.cell(row=tr, column=4).number_format = FMT_PCT1
                # Colorear celda Tendencia con fondo completo
                ws.cell(row=tr, column=5).fill = tend_fill
                ws.cell(row=tr, column=5).font = tend_font
                ws.cell(row=tr, column=5).alignment = _align("center")

        # Anchos
        for i, w in enumerate([22, 16, 16, 14, 18, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Hoja 4: Por Cliente
    # ------------------------------------------------------------------

    def _sheet_clientes(self):
        ws = self.wb.create_sheet("Por Cliente")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:I1")
        ws["A1"] = "Análisis por Cliente — Concentración de Cartera y Segmentación"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = _align("left")
        ws["A1"].fill = _fill("FFF8F0")

        top_clientes = self.proc.get_top_clientes(mode="anual")
        det_df = self.proc.get_detalle_clientes(top_clientes, mode="anual")
        total_global = det_df["Total"].sum() if not det_df.empty else 1

        # ── BLOQUE 1: Concentración de cartera ──────────────────────────
        ws.merge_cells("A3:E3")
        ws["A3"] = "Concentración de Cartera"
        ws["A3"].font = FONT_SECTION
        ws["A3"].fill = FILL_CREAM

        # Calcular concentración
        all_cli = self.proc.get_detalle_clientes(
            self.proc.get_top_clientes(mode="anual"), mode="anual"
        )
        sorted_cli = all_cli.sort_values("Total", ascending=False)
        total_todos = sorted_cli["Total"].sum() or 1
        n_top5  = max(1, int(len(sorted_cli) * 0.05))
        n_top10 = max(1, int(len(sorted_cli) * 0.10))
        pct_top5  = sorted_cli.head(n_top5)["Total"].sum() / total_todos * 100
        pct_top10 = sorted_cli.head(n_top10)["Total"].sum() / total_todos * 100
        n_activos = len(sorted_cli)

        conc_metrics = [
            (f"% Ventas — Top 5% clientes ({n_top5} clientes)",  pct_top5,  True),
            (f"% Ventas — Top 10% clientes ({n_top10} clientes)", pct_top10, True),
            ("# Clientes activos en el período",                   n_activos, False),
        ]
        _header_row(ws, 4, ["Métrica de Concentración", "Valor", "Señal"])
        for ri, (label, valor, es_riesgo) in enumerate(conc_metrics, 5):
            is_num = isinstance(valor, float)
            cell_v = ws.cell(row=ri, column=2, value=valor / 100 if is_num else valor)
            if is_num:
                cell_v.number_format = FMT_PCT1
            ws.cell(row=ri, column=1, value=label).font = FONT_BODY
            ws.cell(row=ri, column=1).alignment = _align("left")
            ws.cell(row=ri, column=1).border = BORDER
            cell_v.border = BORDER
            cell_v.alignment = _align("right")
            # Color riesgo si concentración > 40%
            if es_riesgo and is_num and valor > 40:
                cell_v.font = _font(bold=True, color=TRAFFIC_RED, size=10)
                ws.cell(row=ri, column=3, value="⚠ Concentración alta").font = _font(bold=True, color=TRAFFIC_RED, size=9)
            else:
                cell_v.font = _font(bold=True, color=BRAND_MAROON, size=10)
                ws.cell(row=ri, column=3, value="OK").font = _font(color=TRAFFIC_GREEN, size=9)
            ws.cell(row=ri, column=3).border = BORDER

        # Nota de lectura
        ws.merge_cells("A8:E8")
        ws["A8"] = "Nota: Concentración > 40% en top 10% de clientes representa riesgo de cartera. Diversificar."
        ws["A8"].font = FONT_SUB

        # ── BLOQUE 2: Ranking Top clientes con Pareto ────────────────────
        rank_start = 10
        ws.merge_cells(f"A{rank_start}:J{rank_start}")
        ws[f"A{rank_start}"] = f"Ranking Top {len(top_clientes)} Clientes — Ventas YTD (con Curva Pareto)"
        ws[f"A{rank_start}"].font = FONT_SECTION
        ws[f"A{rank_start}"].fill = FILL_CREAM

        cols = ["#", "Cliente"] + [PRODUCT_DISPLAY_NAMES.get(p, p) for p in PRODUCT_ORDER] + ["Total $", "Part %", "% Acumulado"]
        _header_row(ws, rank_start + 1, cols)

        acumulado = 0.0
        for ri, (idx, row) in enumerate(det_df.iterrows(), rank_start + 2):
            tv    = row["Total"]
            pct   = tv / total_global * 100 if total_global > 0 else 0
            acumulado += pct
            vals  = [ri - (rank_start + 1), str(idx)[:30]]
            vals += [row[p] for p in PRODUCT_ORDER if p in row.index]
            vals += [tv, pct / 100, acumulado / 100]
            _data_row(ws, ri, vals)
            ncols = len(PRODUCT_ORDER)
            for ci in range(3, ncols + 3):
                ws.cell(row=ri, column=ci).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=ncols + 3).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=ncols + 4).number_format = FMT_PCT1
            ws.cell(row=ri, column=ncols + 5).number_format = FMT_PCT1

        # Fila total
        tot_r = rank_start + 2 + len(det_df)
        total_row = ["—", "Total"]
        total_row += [det_df[p].sum() if p in det_df.columns else 0 for p in PRODUCT_ORDER]
        total_row += [total_global, 1.0, 1.0]
        _data_row(ws, tot_r, total_row, bold=True, cream=True)
        ncols = len(PRODUCT_ORDER)
        for ci in range(3, ncols + 4):
            ws.cell(row=tot_r, column=ci).number_format = FMT_CURRENCY

        # ── BLOQUE 3: Clientes en riesgo y en crecimiento ────────────────
        seg = self.proc.get_clientes_riesgo_crecimiento()
        df_riesgo = seg["riesgo"]
        df_crec   = seg["crecimiento"]

        riesgo_start = tot_r + 3
        ws.merge_cells(f"A{riesgo_start}:F{riesgo_start}")
        ws[f"A{riesgo_start}"] = f"Clientes en Riesgo (caída ≥ 20% vs mismo período año anterior) — {len(df_riesgo)} clientes"
        ws[f"A{riesgo_start}"].font = FONT_SECTION
        ws[f"A{riesgo_start}"].fill = _fill("FDECEA")  # fondo rosado suave

        _header_row(ws, riesgo_start + 1,
                    ["Cliente", "Venta Actual $", "Venta Anterior $", "Var $", "Var %"])
        if df_riesgo.empty:
            ws.cell(row=riesgo_start + 2, column=1, value="Sin clientes en riesgo en este período.").font = FONT_SUB
        else:
            for ri2, (idx, row) in enumerate(df_riesgo.iterrows(), riesgo_start + 2):
                _data_row(ws, ri2, [
                    str(idx)[:35],
                    row["venta_actual"],
                    row["venta_anterior"],
                    row["variacion_abs"],
                    row["variacion_pct"] / 100
                ], neg_cols=[4, 5])
                for ci, fmt in [(2, FMT_CURRENCY), (3, FMT_CURRENCY),
                                 (4, FMT_CURRENCY), (5, FMT_PCT1)]:
                    ws.cell(row=ri2, column=ci).number_format = fmt
                ws.cell(row=ri2, column=5).font = FONT_NEG

        crec_section_start = riesgo_start + 3 + max(len(df_riesgo), 1)
        ws.merge_cells(f"A{crec_section_start}:F{crec_section_start}")
        ws[f"A{crec_section_start}"] = f"Clientes en Crecimiento (incremento ≥ 15% vs mismo período año anterior) — {len(df_crec)} clientes"
        ws[f"A{crec_section_start}"].font = FONT_SECTION
        ws[f"A{crec_section_start}"].fill = _fill("E8F5E9")  # fondo verde suave

        _header_row(ws, crec_section_start + 1,
                    ["Cliente", "Venta Actual $", "Venta Anterior $", "Var $", "Var %"])
        if df_crec.empty:
            ws.cell(row=crec_section_start + 2, column=1, value="Sin clientes en crecimiento en este período.").font = FONT_SUB
        else:
            for ri3, (idx, row) in enumerate(df_crec.iterrows(), crec_section_start + 2):
                _data_row(ws, ri3, [
                    str(idx)[:35],
                    row["venta_actual"],
                    row["venta_anterior"],
                    row["variacion_abs"],
                    row["variacion_pct"] / 100
                ])
                for ci, fmt in [(2, FMT_CURRENCY), (3, FMT_CURRENCY),
                                 (4, FMT_CURRENCY), (5, FMT_PCT1)]:
                    ws.cell(row=ri3, column=ci).number_format = fmt
                ws.cell(row=ri3, column=5).font = _font(bold=True, color=TRAFFIC_GREEN, size=9)

        # Anchos
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 32
        for i in range(3, len(PRODUCT_ORDER) + 6):
            ws.column_dimensions[get_column_letter(i)].width = 14

    # ------------------------------------------------------------------
    # Hoja 5: Regional
    # ------------------------------------------------------------------

    def _sheet_regional(self):
        ws = self.wb.create_sheet("Regional")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:H1")
        ws["A1"] = "Análisis Regional — Ventas por Estado (YTD vs Año Anterior)"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = _align("left")
        ws["A1"].fill = _fill("FFF8F0")

        # Obtener año actual y anterior con participaciones
        y, w = self.proc.anio, self.proc.semana
        ly, lyw = y - 1, w

        df_cur = self.proc._filter_ytd(y, w)
        df_ly  = self.proc._filter_ytd(ly, lyw)

        agg_cur = (
            df_cur.groupby("region_o_estado")
            .agg(unidades=("botellas", "sum"), venta_actual=("venta_sin_impuestos", "sum"))
            .reset_index()
        )
        agg_ly = (
            df_ly.groupby("region_o_estado")["venta_sin_impuestos"]
            .sum().reset_index().rename(columns={"venta_sin_impuestos": "venta_anterior"})
        )
        df_reg = agg_cur.merge(agg_ly, on="region_o_estado", how="outer").fillna(0)
        total_cur = df_reg["venta_actual"].sum() or 1
        total_ly  = df_reg["venta_anterior"].sum() or 1
        df_reg["pct_actual"]   = df_reg["venta_actual"]   / total_cur * 100
        df_reg["pct_anterior"] = df_reg["venta_anterior"] / total_ly  * 100
        df_reg["var_pp"]       = df_reg["pct_actual"] - df_reg["pct_anterior"]
        df_reg = df_reg.sort_values("venta_actual", ascending=False).reset_index(drop=True)

        ws.merge_cells("A3:H3")
        ws["A3"] = f"Ventas por Estado — YTD {y} vs {ly} (hasta S{w:02d})"
        ws["A3"].font = FONT_SECTION
        ws["A3"].fill = FILL_CREAM

        _header_row(ws, 4, [
            "#", "Estado / Región",
            f"Venta {y} $", f"Venta {ly} $",
            f"Part% {y}", f"Part% {ly}",
            "Var pp"
        ])

        last_data_row = 4 + len(df_reg)
        for ri, (_, row) in enumerate(df_reg.iterrows(), 5):
            var_pp = row["var_pp"]
            _data_row(ws, ri, [
                ri - 4,
                row["region_o_estado"],
                row["venta_actual"],
                row["venta_anterior"],
                row["pct_actual"]   / 100,
                row["pct_anterior"] / 100,
                var_pp / 100
            ])
            ws.cell(row=ri, column=3).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=4).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=5).number_format = FMT_PCT1
            ws.cell(row=ri, column=6).number_format = FMT_PCT1
            ws.cell(row=ri, column=7).number_format = "0.0pp"
            # Colorear variación pp: verde si ganó participación, rojo si perdió
            pp_cell = ws.cell(row=ri, column=7)
            if var_pp > 0:
                pp_cell.font = _font(bold=True, color=TRAFFIC_GREEN, size=9)
            elif var_pp < 0:
                pp_cell.font = _font(bold=True, color=TRAFFIC_RED, size=9)

        # DataBar en ventas año actual
        ws.conditional_formatting.add(
            f"C5:C{last_data_row}",
            DataBarRule(start_type="min", end_type="max",
                        color=BRAND_MAROON.lstrip("#"))
        )

        # Gráfica barras horizontales (top 15) — actual vs anterior
        chart = BarChart()
        chart.type = "bar"  # horizontal
        chart.grouping = "clustered"
        chart.title = f"Ventas por Estado: {y} vs {ly}"
        chart.x_axis.title = "Estado"          # TextAxis (categorías / estados)
        chart.y_axis.title = "$MXN sin IVA"    # NumericAxis (valores)
        chart.width = 24
        chart.height = 15
        chart.style = 10

        n_rows = min(15, len(df_reg))
        data = Reference(ws, min_col=3, max_col=4, min_row=4, max_row=4 + n_rows)
        cats = Reference(ws, min_col=2, min_row=5, max_row=4 + n_rows)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Ordenar para que el estado #1 aparezca arriba y etiquetas visibles
        chart.x_axis.scaling.orientation = "maxMin"
        chart.y_axis.crosses = "min"
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"

        # Colores de marca Design.md: Año actual Maroon #6E1E28, Año anterior Khaki #E7D6A6
        if len(chart.series) > 0:
            chart.series[0].graphicalProperties.solidFill = "6E1E28"
        if len(chart.series) > 1:
            chart.series[1].graphicalProperties.solidFill = "E7D6A6"

        _add_chart_with_note(ws, chart, "I3")

        for i, w_col in enumerate([5, 28, 18, 18, 12, 12, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w_col

    # ------------------------------------------------------------------
    # Hoja 6: Por Canal (nueva hoja separada)
    # ------------------------------------------------------------------

    def _sheet_canal(self):
        ws = self.wb.create_sheet("Por Canal")
        ws.sheet_view.showGridLines = False

        y, w = self.proc.anio, self.proc.semana
        ly = y - 1

        ws.merge_cells("A1:H1")
        ws["A1"] = f"Análisis por Canal — YTD {y} vs {ly} (hasta S{w:02d})"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = _align("left")
        ws["A1"].fill = _fill("FFF8F0")

        df_canal = self.proc.get_canal_analisis(mode="anual")

        ws.merge_cells("A3:H3")
        ws["A3"] = f"Ventas por Canal de Distribución — {y} vs {ly}"
        ws["A3"].font = FONT_SECTION
        ws["A3"].fill = FILL_CREAM

        _header_row(ws, 4, [
            "Canal",
            f"Unidades {y}",
            f"Venta {y} $",
            f"Venta {ly} $",
            f"Part% {y}",
            f"Part% {ly}",
            "Var pp"
        ])

        for ri, (_, row) in enumerate(df_canal.iterrows(), 5):
            var_pp = row["var_pp"]
            _data_row(ws, ri, [
                row["canal"],
                row["unidades"],
                row["venta_actual"],
                row["venta_anterior"],
                row["pct_part_actual"]   / 100,
                row["pct_part_anterior"] / 100,
                var_pp / 100
            ])
            ws.cell(row=ri, column=2).number_format = FMT_INT
            ws.cell(row=ri, column=3).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=4).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=5).number_format = FMT_PCT1
            ws.cell(row=ri, column=6).number_format = FMT_PCT1
            ws.cell(row=ri, column=7).number_format = "0.0pp"
            # Variación pp de canal: sin color (consistente con archivo original)

        # Fila total
        tot_r = 5 + len(df_canal)
        _data_row(ws, tot_r, [
            "Total",
            df_canal["unidades"].sum(),
            df_canal["venta_actual"].sum(),
            df_canal["venta_anterior"].sum(),
            1.0, 1.0, 0.0
        ], bold=True, cream=True)
        ws.cell(row=tot_r, column=2).number_format = FMT_INT
        ws.cell(row=tot_r, column=3).number_format = FMT_CURRENCY
        ws.cell(row=tot_r, column=4).number_format = FMT_CURRENCY
        ws.cell(row=tot_r, column=5).number_format = FMT_PCT1
        ws.cell(row=tot_r, column=6).number_format = FMT_PCT1

        # ── Gráfica: Barras agrupadas por canal (año actual vs anterior) ──────
        try:
            chart = BarChart()
            chart.type = "bar"
            chart.grouping = "clustered"
            chart.title = f"Ventas por Canal: {y} vs {ly}"
            chart.x_axis.title = "Canal"            # TextAxis (Categorías / Canales)
            chart.y_axis.title = "$MXN sin IVA"      # NumericAxis (Valores en Pesos)
            chart.x_axis.scaling.orientation = "maxMin"
            chart.y_axis.crosses = "min"
            chart.x_axis.tickLblPos = "nextTo"
            chart.y_axis.tickLblPos = "nextTo"
            chart.width = 24
            chart.height = 14
            chart.style = 10

            n_canales = len(df_canal)
            data = Reference(ws, min_col=3, max_col=4,
                             min_row=4, max_row=4 + n_canales)
            cats = Reference(ws, min_col=1, min_row=5, max_row=4 + n_canales)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # Colores de marca Design.md: Año actual Maroon #6E1E28, Año anterior Khaki #E7D6A6
            if len(chart.series) > 0:
                chart.series[0].graphicalProperties.solidFill = "6E1E28"
            if len(chart.series) > 1:
                chart.series[1].graphicalProperties.solidFill = "E7D6A6"

            _add_chart_with_note(ws, chart, "I3")
        except Exception:
            pass

        # ── Tabla secundaria simplificada — alimenta la gráfica Pareto ───
        row_pareto = tot_r + 3
        ws.merge_cells(f"A{row_pareto}:D{row_pareto}")
        ws[f"A{row_pareto}"] = "Mix de Canal — Tabla Simplificada"
        ws[f"A{row_pareto}"].font = FONT_SECTION
        ws[f"A{row_pareto}"].fill = FILL_CREAM
        _header_row(ws, row_pareto + 1, ["Canal", f"Venta {y} $", f"Part% {y}"])
        for ri2, (_, row) in enumerate(df_canal.iterrows(), row_pareto + 2):
            _data_row(ws, ri2, [
                row["canal"],
                row["venta_actual"],
                row["pct_part_actual"] / 100
            ])
            ws.cell(row=ri2, column=2).number_format = FMT_CURRENCY
            ws.cell(row=ri2, column=3).number_format = FMT_PCT1

        for i, w_col in enumerate([28, 14, 18, 18, 12, 12, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w_col

    # ------------------------------------------------------------------
    # Hoja 7: Oportunidades y Riesgos
    # ------------------------------------------------------------------

    def _sheet_oportunidades(self):
        ws = self.wb.create_sheet("Oportunidades y Riesgos")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:E1")
        ws["A1"] = "Oportunidades y Riesgos Identificados"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = _align("left")
        ws["A1"].fill = _fill("FFF8F0")

        ws.merge_cells("A2:E2")
        ws["A2"] = (f"Análisis automatizado · Semana {self.proc.semana:02d} · {self.proc.anio} "
                    "— Complementar con contexto de mercado")
        ws["A2"].font = FONT_SUB

        _header_row(ws, 4, ["#", "Hallazgo", "Tipo", "Impacto Estimado", "Recomendación de Acción"])

        items = self.proc.get_oportunidades_riesgos()
        for ri, item in enumerate(items, 5):
            # Color de fondo COMPLETO en columna Tipo (fondo + texto blanco)
            tipo = item["tipo"]
            if tipo == "Riesgo":
                tipo_fill = _fill("A6192E")
                tipo_font = _font(bold=True, color="FFFFFF", size=9)
            elif tipo == "Oportunidad":
                tipo_fill = _fill("1E7145")
                tipo_font = _font(bold=True, color="FFFFFF", size=9)
            elif tipo in ("Riesgo / Oportunidad", "Riesgo-Oportunidad"):
                tipo_fill = _fill("8A6D00")
                tipo_font = _font(bold=True, color="FFFFFF", size=9)
            else:
                tipo_fill = FILL_HEADER2
                tipo_font = FONT_BODY

            vals = [ri - 4, item["hallazgo"], item["tipo"], item["impacto"], item["recomendacion"]]
            _data_row(ws, ri, vals)
            ws.cell(row=ri, column=3).fill = tipo_fill
            ws.cell(row=ri, column=3).font = tipo_font
            ws.cell(row=ri, column=3).alignment = _align("center")
            for ci in range(1, 6):
                ws.cell(row=ri, column=ci).alignment = _align("left", wrap=True)
            ws.cell(row=ri, column=3).alignment = _align("center", wrap=False)
            ws.row_dimensions[ri].height = 28

        # Conclusiones
        conc_row = 5 + len(items) + 2
        ws.merge_cells(f"A{conc_row}:E{conc_row}")
        ws[f"A{conc_row}"] = "Conclusiones y Próximos Pasos"
        ws[f"A{conc_row}"].font = FONT_SECTION
        ws[f"A{conc_row}"].fill = FILL_CREAM

        resumen = self.proc.get_resumen_ejecutivo()
        wow_pct = resumen["vs_semana_anterior"]["pct"]
        ytd_pct = resumen["ytd_vs_ly"]["pct"]
        plan_pct = resumen["vs_plan"]["pct"]

        pasos = [
            f"1. WoW: {wow_pct:+.1f}% — {'Mantener ritmo' if wow_pct > 0 else 'Investigar causa de caída'}",
            f"2. YTD vs Año Ant.: {ytd_pct:+.1f}% — {'Sostener estrategia actual' if ytd_pct > 0 else 'Revisar mix de producto y canal'}",
            f"3. Cumplimiento de plan: {plan_pct:+.1f}% — {'OK' if plan_pct > -5 else 'Requiere plan de recuperación urgente'}",
            "4. Monitorear concentración de clientes — diversificar cartera si top 3 > 50%.",
            "5. Validar datos de mercado (CRT, exportaciones, precio agave) para contexto externo.",
        ]
        for i, paso in enumerate(pasos, conc_row + 1):
            ws.merge_cells(f"A{i}:E{i}")
            ws[f"A{i}"] = paso
            ws[f"A{i}"].font = FONT_BODY
            ws[f"A{i}"].alignment = _align("left", wrap=True)
            ws.row_dimensions[i].height = 20

        # Anchos
        widths = [5, 40, 16, 24, 45]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Hoja 7 (opcional): Comparativo Custom entre dos periodos
    # ------------------------------------------------------------------

    def _sheet_comparativo_custom(self, comp: Dict):
        """
        Hoja generada solo cuando se pasa --comparar-*.
        Muestra lado a lado los dos periodos con variaciones.
        """
        ws = self.wb.create_sheet("Comparativo Custom")
        ws.sheet_view.showGridLines = False

        label_a = comp["labels"]["a"]
        label_b = comp["labels"]["b"]

        ws.merge_cells("A1:G1")
        ws["A1"] = f"Comparativo: {label_a}  vs  {label_b}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = _align("left")
        ws["A1"].fill = _fill("FFF8F0")

        # --- KPIs comparativos ---
        ws.merge_cells("A3:G3")
        ws["A3"] = "KPIs Generales"
        ws["A3"].font = FONT_SECTION
        ws["A3"].fill = FILL_CREAM

        _header_row(ws, 4, ["Metrica", label_a, label_b,
                              "Variacion $", "Variacion %", "Tendencia"])

        metricas = [
            ("Ventas Netas $",
             comp["kpis_a"]["ventas_netas"],
             comp["kpis_b"]["ventas_netas"],
             comp["variacion"]["ventas_netas"]["abs"],
             comp["variacion"]["ventas_netas"]["pct"]),
            ("# Botellas",
             comp["kpis_a"]["botellas"],
             comp["kpis_b"]["botellas"],
             comp["variacion"]["botellas"]["abs"],
             comp["variacion"]["botellas"]["pct"]),
            ("Margen %",
             comp["kpis_a"]["margen_pct"],
             comp["kpis_b"]["margen_pct"],
             comp["variacion"]["margen_pct"]["abs"],
             comp["variacion"]["margen_pct"]["pct"]),
            ("Ticket Promedio $",
             comp["kpis_a"]["ticket_promedio"],
             comp["kpis_b"]["ticket_promedio"],
             comp["variacion"]["ticket_promedio"]["abs"],
             comp["variacion"]["ticket_promedio"]["pct"]),
        ]

        for ri, (met, va, vb, var_abs, var_pct) in enumerate(metricas, 5):
            cream = (ri == 5)
            tendencia = "mejora" if var_abs >= 0 else "baja"
            tend_color = TRAFFIC_GREEN if var_abs >= 0 else TRAFFIC_RED
            _data_row(ws, ri, [met, va, vb, var_abs, var_pct / 100, tendencia],
                      cream=cream, neg_cols=[4])
            ws.cell(row=ri, column=2).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=3).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=4).number_format = FMT_CURRENCY
            ws.cell(row=ri, column=5).number_format = FMT_PCT1
            ws.cell(row=ri, column=6).font = _font(bold=True, color=tend_color, size=9)

        # --- Por Producto ---
        row = 11
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = "Comparativo por Producto"
        ws[f"A{row}"].font = FONT_SECTION
        ws[f"A{row}"].fill = FILL_CREAM

        por_prod = comp.get("por_producto")
        if por_prod is not None and not por_prod.empty:
            _header_row(ws, row + 1, ["Producto", label_a, label_b,
                                       "Variacion $", "Variacion %"])
            for ri2, (idx, prow) in enumerate(por_prod.iterrows(), row + 2):
                va = prow.get("periodo_a", 0)
                vb = prow.get("periodo_b", 0)
                vd = prow.get("variacion_abs", 0)
                vp = prow.get("variacion_pct", 0)
                prod_display = PRODUCT_DISPLAY_NAMES.get(idx, idx)
                _data_row(ws, ri2, [prod_display, va, vb, vd, vp / 100],
                          neg_cols=[4])
                for ci, fmt in [(2, FMT_CURRENCY), (3, FMT_CURRENCY),
                                 (4, FMT_CURRENCY), (5, FMT_PCT1)]:
                    ws.cell(row=ri2, column=ci).number_format = fmt

            # Grafica de barras agrupadas
            last_prod_row = row + 1 + len(por_prod)
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = f"Ventas por Producto: {label_a} vs {label_b}"
            chart.y_axis.title = "$MXN sin IVA"
            chart.x_axis.tickLblPos = "nextTo"
            chart.y_axis.tickLblPos = "nextTo"
            chart.width = 22
            chart.height = 12
            chart.style = 10

            data_ref = Reference(ws, min_col=2, max_col=3,
                                  min_row=row + 1, max_row=last_prod_row)
            cats_ref = Reference(ws, min_col=1,
                                  min_row=row + 2, max_row=last_prod_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            _add_chart_with_note(ws, chart, f"G{row}")
            next_section = last_prod_row + 16
        else:
            next_section = row + 4

        # --- Por Canal ---
        row2 = next_section
        ws.merge_cells(f"A{row2}:G{row2}")
        ws[f"A{row2}"] = "Comparativo por Canal"
        ws[f"A{row2}"].font = FONT_SECTION
        ws[f"A{row2}"].fill = FILL_CREAM

        por_canal = comp.get("por_canal")
        if por_canal is not None and not por_canal.empty:
            _header_row(ws, row2 + 1, ["Canal", label_a, label_b,
                                         "Variacion $", "Variacion %"])
            for ri3, (idx, crow) in enumerate(por_canal.iterrows(), row2 + 2):
                va = crow.get("periodo_a", 0)
                vb = crow.get("periodo_b", 0)
                vd = crow.get("variacion_abs", 0)
                vp = crow.get("variacion_pct", 0)
                _data_row(ws, ri3, [idx, va, vb, vd, vp / 100], neg_cols=[4])
                for ci, fmt in [(2, FMT_CURRENCY), (3, FMT_CURRENCY),
                                 (4, FMT_CURRENCY), (5, FMT_PCT1)]:
                    ws.cell(row=ri3, column=ci).number_format = fmt

        # Anchos
        for i, w in enumerate([22, 18, 18, 16, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Hoja 8 (opcional): Contexto de mercado externo
    # ------------------------------------------------------------------

    def _sheet_contexto_mercado(self, contexto: Dict):
        """Hoja con el contexto de mercado (CRT, agave, NOM) buscado por el agente."""
        ws = self.wb.create_sheet("Contexto Mercado")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:E1")
        ws["A1"] = "Contexto de Mercado Externo"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = _fill("FFF8F0")

        if not contexto.get("disponible", False):
            ws.merge_cells("A3:E3")
            ws["A3"] = contexto.get("resumen", "Sin contexto de mercado disponible.")
            ws["A3"].font = FONT_SUB
            ws["A3"].alignment = _align("left", wrap=True)
            ws.row_dimensions[3].height = 40
            ws.merge_cells("A5:E5")
            ws["A5"] = ("Para incluir contexto de mercado, solicitar al agente: "
                        "'busca estadisticas CRT, precio agave y cambios NOM-006 "
                        "para la semana {semana} de {anio} y agrega el contexto al reporte'")
            ws["A5"].font = _font(italic=True, color=SECTION_SUBTITLE, size=9)
            ws["A5"].alignment = _align("left", wrap=True)
            ws.row_dimensions[5].height = 40
        else:
            # Resumen general
            ws.merge_cells("A3:E5")
            ws["A3"] = contexto["resumen"][:2000]
            ws["A3"].alignment = _align("left", wrap=True)
            ws["A3"].font = FONT_BODY
            ws.row_dimensions[3].height = 80

            # Hallazgos del mercado
            row = 7
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"] = "Hallazgos de Mercado"
            ws[f"A{row}"].font = FONT_SECTION
            ws[f"A{row}"].fill = FILL_CREAM

            _header_row(ws, row + 1, ["#", "Hallazgo", "Tipo", "Impacto", "Recomendacion"])
            for ri, h in enumerate(contexto.get("hallazgos", []), row + 2):
                tipo_color = TRAFFIC_RED if h["tipo"] == "Riesgo" else (
                    TRAFFIC_GREEN if h["tipo"] == "Oportunidad" else TRAFFIC_YELLOW)
                _data_row(ws, ri, [ri - (row + 1), h["hallazgo"], h["tipo"],
                                    h["impacto"], h["recomendacion"]])
                ws.cell(row=ri, column=3).font = _font(bold=True, color=tipo_color, size=9)

            # Fuentes
            if contexto.get("fuentes"):
                src_row = row + 2 + len(contexto["hallazgos"]) + 2
                ws.merge_cells(f"A{src_row}:E{src_row}")
                ws[f"A{src_row}"] = "Fuentes consultadas:"
                ws[f"A{src_row}"].font = FONT_SECTION
                ws[f"A{src_row}"].fill = FILL_CREAM

                for idx_src, fuente in enumerate(contexto["fuentes"], src_row + 1):
                    ws.merge_cells(f"A{idx_src}:E{idx_src}")
                    ws[f"A{idx_src}"] = f"•  {fuente}"
                    ws[f"A{idx_src}"].font = FONT_BODY
                    ws[f"A{idx_src}"].alignment = _align("left", wrap=True)
                    ws.row_dimensions[idx_src].height = 20

        for i, w in enumerate([5, 45, 14, 24, 40], 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Funcion de entrada
# ---------------------------------------------------------------------------

def generate_xlsx(
    processor: LocoDataProcessor,
    output_path: str,
    comparativo_custom: Optional[Dict] = None,
    contexto_mercado: Optional[Dict] = None,
):
    """
    Genera el XLSX ejecutivo y lo guarda en output_path.

    Params opcionales:
      comparativo_custom : resultado de proc.get_comparativo_custom(...)
      contexto_mercado   : resultado de market_context.load_market_context(...)
    """
    gen = LocoReporteXLSX(processor, output_path)

    # Hojas base siempre
    gen._sheet_resumen()
    gen._sheet_comparativo()
    gen._sheet_productos()
    gen._sheet_clientes()
    gen._sheet_regional()
    gen._sheet_canal()
    gen._sheet_oportunidades()

    # Hojas opcionales
    if comparativo_custom:
        gen._sheet_comparativo_custom(comparativo_custom)

    if contexto_mercado:
        gen._sheet_contexto_mercado(contexto_mercado)
    else:
        # Siempre incluir hoja de contexto, aunque sea el placeholder
        gen._sheet_contexto_mercado({"disponible": False,
                                      "resumen": "", "fuentes": [], "hallazgos": []})

    try:
        gen.wb.save(output_path)
        print(f"[XLSX] Guardado: {output_path}")
    except PermissionError:
        alt_path = output_path.replace(".xlsx", "_actualizado.xlsx")
        gen.wb.save(alt_path)
        print(f"[XLSX] ⚠️ El archivo original está abierto en Excel. Guardado copia en: {alt_path}")
